import pandas as pd
from flask import send_file, jsonify
import pandas as pd
import networkx as nx
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
from flask_cors import CORS


def process_bom_analysis(request):
    try:
        bom_file = request.files['bom_file']
        central_bom_id = request.form['central_bom_id']
        max_distance = int(request.form['max_distance'])

        df = pd.read_csv(bom_file)

        # Convert BOM IDs to strings to ensure consistency
        df['BOM ID'] = df['BOM ID'].astype(str)
        df['Next BOM ID'] = df['Next BOM ID'].astype(str)

        G = nx.DiGraph()
        for _, row in df.iterrows():
            G.add_edge(str(row['BOM ID']), str(row['Next BOM ID']))

        # Ensure central_bom_id is a string
        central_bom_id = str(central_bom_id)

        if central_bom_id not in G:
            return jsonify({'error': f"Central BOM ID {central_bom_id} not found in the graph"}), 400

        precursors, successors = find_precursors_and_successors(G, central_bom_id, max_distance)

        output = write_to_excel(df, central_bom_id, max_distance, precursors, successors)

        filename = f"BOM_{central_bom_id}_{max_distance}_edges.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except KeyError as e:
        return jsonify({'error': f"Missing required field: {str(e)}"}), 400
    except ValueError as e:
        return jsonify({'error': f"Invalid value: {str(e)}"}), 400
    except Exception as e:
        print(f"Error in process_bom_analysis: {str(e)}")
        return jsonify({'error': str(e)}), 500

def find_precursors_and_successors(G, central_node, max_distance):
    central_node = str(central_node)  # Ensure central_node is a string
    precursors = defaultdict(list)
    successors = defaultdict(list)

    # Find Precursors
    queue = [(central_node, 1)]
    visited = set()
    while queue:
        node, distance = queue.pop()
        if distance > max_distance:
            continue
        for predecessor in G.predecessors(node):
            if predecessor not in visited:
                precursors[distance].append(predecessor)
                if distance < max_distance:
                    queue.append((predecessor, distance + 1))
                visited.add(predecessor)

    # Now find Successors
    queue = [(central_node, 1)]
    visited = set()
    while queue:
        node, distance = queue.pop(0)
        if distance > max_distance:
            continue
        for successor in G.successors(node):
            if successor not in visited:
                successors[distance].append(successor)
                if distance < max_distance:
                    queue.append((successor, distance + 1))
                visited.add(successor)

    return precursors, successors

def write_to_excel(df, central_bom_id, max_distance, precursors, successors):
    wb = Workbook()
    ws = wb.active
    ws.title = f"BOM {central_bom_id} - {max_distance} edges"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    centered_alignment = Alignment(horizontal='center', vertical='center')

    # Write headers
    headers = ['Distance', 'Precursor BOM ID', 'Precursor Description', 'Successor BOM ID', 'Successor Description']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment

    # Write data
    row = 2
    for distance in range(max_distance + 1):
        precursor_nodes = precursors.get(distance, [])
        successor_nodes = successors.get(distance, [])
        max_nodes = max(len(precursor_nodes), len(successor_nodes))

        for i in range(max_nodes):
            ws.cell(row=row, column=1, value=distance)

            if i < len(precursor_nodes):
                precursor = precursor_nodes[i]
                ws.cell(row=row, column=2, value=precursor)
                ws.cell(row=row, column=3, value=get_description(df, precursor))

            if i < len(successor_nodes):
                successor = successor_nodes[i]
                ws.cell(row=row, column=4, value=successor)
                ws.cell(row=row, column=5, value=get_next_description(df, successor))

            row += 1

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def get_description(df, bom_id):
    desc = df[df['BOM ID'] == bom_id]['Description'].values
    return desc[0] if len(desc) > 0 else 'N/A'

def get_next_description(df, bom_id):
    desc = df[df['Next BOM ID'] == bom_id]['Description.1'].values
    return desc[0] if len(desc) > 0 else 'N/A'